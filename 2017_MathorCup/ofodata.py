#!/usr/bin/python
# -*- coding: UTF-8 -*-

# 本程序将B题附件1的txt文件转换为特定排列格式的xlsx表格文件
# 编程环境：Python 3.5.2

import re
import openpyxl

# 创建新表格
print('创建表格...')
wb = openpyxl.Workbook()
ws = wb.active
ws['A1'] = '出发时间'
ws['B1'] = '出发区域'
ws['C1'] = '到达时间'
ws['D1'] = '到达区域'
ws['E1'] = '自行车序号'
print('创建表格完毕！')
print('读取数据...')
f = open('附件1 骑行数据.txt', 'r')
print('读取数据完毕！')
lines = f.readlines()
num = 1 # 自行车序号
departure = 8 # 出发区域
count = 2 # 表格行数
print('写入数据到表格中...')
for line in lines:
    # 将每一行的数据分割为多个字符串
    items = re.sub('\n', '', line)
    items = re.split(',|:' ,items)
    # print(items)
    # 分别将每一辆自行车的所有行驶数据录入表格
    if items[0] == '自行车序号':
        num = int(items[1])
        departure = int(items[3])
    elif items != ['']:
        ws.cell(row=count, column=1).value = float(items[1]) # 出发时间
        ws.cell(row=count, column=2).value = departure # 出发区域
        ws.cell(row=count, column=3).value = float(items[3]) # 到达时间
        ws.cell(row=count, column=4).value = int(items[5]) # 到达区域
        ws.cell(row=count, column=5).value = num # 自行车序号
        departure = int(items[5])
        count = count + 1
print('写入数据完毕！')
print('保存表格...')
wb.save('ofodata.xlsx')
print('Done!')