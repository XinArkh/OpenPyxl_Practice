#!/usr/bin/env python
# -*- coding: UTF-8 -*-

from openpyxl import load_workbook
from urllib.request import urlopen
from urllib.parse import quote
import hashlib
import json
import random


class Translation(object):
    __appKey = 'MYAPPKEY'
    __secretKey = 'MYSECRETKEY'
    myURL = 'http://openapi.youdao.com/api'
    langList = {
        'auto': 'auto',
        '中文': 'zh-CHS',
        '日文': 'ja',
        '英文': 'EN',
        '韩文': 'ko',
        '法文': 'fr',
        '俄文': 'ru',
        '葡文': 'pt',
        '西文': 'es'}

    def __init__(self, q, fromLang='auto', toLang='auto'):
        self.fromLang = self.langList[fromLang]
        self.toLang = self.langList[toLang]
        salt = random.randint(1, 65536)
        sign = self.__appKey + q + str(salt) + self.__secretKey
        sign = hashlib.md5(sign.encode(encoding='utf-8')).hexdigest()
        self.myURL = self.myURL + '?appKey=' + self.__appKey + '&q=' + quote(q) + '&from=' + self.fromLang +\
            '&to=' + self.toLang + '&salt=' + str(salt) + '&sign=' + sign
        # print(self.myURL)

    def translate(self):
        try:
            self.res = urlopen(self.myURL).read().decode('UTF-8')
            # print('here', self.res)
        except Exception as e:
            print(e)
            return
        self.resJson = json.loads(self.res)
        # print(self.resJson)
        if self.resJson.get('errorCode') == '0':
            return self.resJson
        else:
            print(
                'Error in mod_translation.py-->class Translation-->translate: %s' %
                self.resJson.get('errorCode'))
            return

    def __call__(self):
        # print(self.translate())
        trs = self.translate()['translation'][0]
        return trs


wb = load_workbook('ProblemCData.xlsx')
ws = wb.get_sheet_by_name('msncodes')
colB = ws['B']
for rx in range(2, ws.max_row + 1):
    cell = ws.cell(row=rx, column=2).value
    f = Translation(cell)
    ws.cell(row=rx, column=4).value = f()

wb.save('ProblemCData-translated.xlsx')
