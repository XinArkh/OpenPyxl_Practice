#!/usr/bin/env python
# -*- coding: UTF-8 -*-

from openpyxl import load_workbook

wb = load_workbook(filename='ProblemCData.xlsx')
ws = wb.get_sheet_by_name('msncodes')
colC = ws['C']
units = []
for rx in range(2, ws.max_row + 1):
    cell = ws.cell(row=rx, column=3).value
    if cell not in units:
        units.append(cell)

print('Units:\n')
for i in range(len(units)):
    print('%2d: ' % (i + 1) + str(units[i]))
